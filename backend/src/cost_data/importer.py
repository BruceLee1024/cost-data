from __future__ import annotations

import hashlib
import json
import shutil
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from cost_data.config import get_settings
from cost_data.db import SessionLocal
from cost_data.fixedpoint import to_scaled
from cost_data.models import (
    CostItem,
    FeeRate,
    ImportIssue,
    ImportJob,
    MeasureItem,
    ProjectVersion,
    ParserProfile,
    QuotaItem,
    RateComponent,
    ResourceItem,
    SourceFile,
)
from cost_data.normalization import normalize_text, normalize_unit


HEADER_ALIASES = {
    "code": {"清单编码", "项目编码", "编码", "定额编号", "材料编码"},
    "name": {"项目名称", "清单名称", "名称", "材料名称", "费用名称", "定额名称"},
    "description": {"项目特征", "项目特征描述", "工作内容", "描述"},
    "specification": {"规格型号", "规格", "型号", "特征"},
    "unit": {"计量单位", "单位"},
    "quantity": {"工程量", "默认工程量", "数量", "消耗量", "含量"},
    "unit_price": {"综合单价", "市场价", "单价", "除税单价"},
    "total": {"合价", "综合合价", "金额", "费用"},
    "category": {"费用类别", "类别", "构成", "费用构成", "资源类别", "工料机类别"},
    "basis": {"取费基础", "计算基础", "取费基数"},
    "rate": {"费率", "取费费率"},
    "specialty": {"专业"},
    "applicable_scope": {"适用范围"},
    "structure_group": {"结构分组"},
    "related_quota_codes": {"关联定额编码"},
    "source": {"来源"},
    "version": {"版本"},
    "remark": {"备注"},
}


@dataclass
class ParsedTable:
    report_type: str
    sheet_name: str
    header_row: int
    columns: dict[str, int]
    header_rows: list[int] = field(default_factory=list)
    header_paths: dict[str, list[str]] = field(default_factory=dict)
    all_header_paths: dict[int, list[str]] = field(default_factory=dict)


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def archive_file(source: Path, original_name: str) -> tuple[str, str, int]:
    settings = get_settings()
    digest = file_sha256(source)
    suffix = Path(original_name).suffix.lower()
    relative = Path(digest[:2]) / f"{digest}{suffix}"
    destination = settings.raw_dir / relative
    destination.parent.mkdir(parents=True, exist_ok=True)
    if not destination.exists():
        shutil.copy2(source, destination)
    return digest, str(relative), destination.stat().st_size


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_columns(row: tuple[Cell, ...]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for index, cell in enumerate(row, start=1):
        value = normalize_text(_cell_text(cell.value)).replace(" ", "")
        if not value:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if value in {normalize_text(alias).replace(" ", "") for alias in aliases}:
                columns.setdefault(field, index)
    return columns


def _classify(sheet_name: str, columns: dict[str, int]) -> str:
    name = normalize_text(sheet_name)
    # A resource category column is common in 工料机汇总表.  Sheet intent takes
    # precedence over the generic category/code/amount shape of a rate analysis.
    if "工料机" in name or "材料" in name or "人工" in name or "机械" in name:
        return "resource"
    if "综合单价分析" in name or ("category" in columns and "code" in columns and "total" in columns):
        return "rate_analysis"
    if "措施" in name:
        return "measure"
    if "取费" in name or "费率" in name or "rate" in columns:
        return "fee_rate"
    if "定额" in name and "quantity" in columns:
        return "quota"
    if "name" in columns and ("unit_price" in columns or "total" in columns):
        return "bill"
    if "name" in columns and "code" in columns and "unit" in columns:
        return "catalog_bill"
    return "unknown"


def _merged_anchor(worksheet: Any) -> dict[tuple[int, int], tuple[int, int]]:
    anchors: dict[tuple[int, int], tuple[int, int]] = {}
    for merged_range in worksheet.merged_cells.ranges:
        anchor = (merged_range.min_row, merged_range.min_col)
        for row_no in range(merged_range.min_row, merged_range.max_row + 1):
            for column_no in range(merged_range.min_col, merged_range.max_col + 1):
                anchors[(row_no, column_no)] = anchor
    return anchors


def _header_value(worksheet: Any, anchors: dict[tuple[int, int], tuple[int, int]], row_no: int, column_no: int) -> str:
    anchor = anchors.get((row_no, column_no), (row_no, column_no))
    return _cell_text(worksheet.cell(*anchor).value)


def _header_rows(worksheet: Any, header_row: int, anchors: dict[tuple[int, int], tuple[int, int]]) -> list[int]:
    rows = [header_row]
    for row_no in range(header_row - 1, max(header_row - 5, 0), -1):
        if not any(_header_value(worksheet, anchors, row_no, column_no) for column_no in range(1, min(worksheet.max_column, 200) + 1)):
            break
        rows.insert(0, row_no)
    return rows


def _inspect_sheet(worksheet: Any) -> ParsedTable | None:
    anchors = _merged_anchor(worksheet)
    for row_no in range(1, min(worksheet.max_row, 80) + 1):
        row = tuple(worksheet.cell(row_no, column_no) for column_no in range(1, min(worksheet.max_column, 200) + 1))
        columns = _find_columns(row)
        if "name" not in columns or len(columns) < 2:
            continue
        header_rows = _header_rows(worksheet, row_no, anchors)
        header_paths: dict[str, list[str]] = {}
        all_header_paths: dict[int, list[str]] = {}
        for column_no in range(1, min(worksheet.max_column, 200) + 1):
            path = [
                value
                for header_row in header_rows
                if (value := _header_value(worksheet, anchors, header_row, column_no))
            ]
            if path:
                all_header_paths[column_no] = path
        for field, column_no in columns.items():
            header_paths[field] = all_header_paths.get(column_no, [])
        return ParsedTable(
            report_type=_classify(worksheet.title, columns),
            sheet_name=worksheet.title,
            header_row=row_no,
            columns=columns,
            header_rows=header_rows,
            header_paths=header_paths,
            all_header_paths=all_header_paths,
        )
    return None


def inspect_workbook(path: Path) -> tuple[list[str], list[ParsedTable]]:
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    tables: list[ParsedTable] = []
    sheet_names = list(workbook.sheetnames)
    try:
        for worksheet in workbook.worksheets:
            table = _inspect_sheet(worksheet)
            if table:
                tables.append(table)
    finally:
        workbook.close()
    return sheet_names, tables


def _value(row: tuple[Any, ...], columns: dict[str, int], key: str) -> Any:
    index = columns.get(key)
    return row[index - 1] if index and index <= len(row) else None


def _table_preview(source_file: SourceFile, table: ParsedTable) -> dict[str, Any]:
    fingerprint = hashlib.sha256(
        json.dumps(
            {
                "sheet_name": normalize_text(table.sheet_name),
                "report_type": table.report_type,
                "header_paths": table.header_paths,
                "all_header_paths": table.all_header_paths,
            },
            ensure_ascii=False,
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()
    return {
        "source_file_id": source_file.id,
        "sheet_name": table.sheet_name,
        "report_type": table.report_type,
        "fingerprint": fingerprint,
        "header_rows": table.header_rows or [table.header_row],
        "columns": {
            field: {"column": column_no, "header_path": table.header_paths.get(field, [])}
            for field, column_no in table.columns.items()
        },
        "raw_columns": {str(column_no): path for column_no, path in table.all_header_paths.items()},
        "requires_confirmation": len(table.header_rows or [table.header_row]) > 1 or table.report_type == "unknown",
    }


def save_parser_profiles(session: Session, tables: list[dict[str, Any]]) -> None:
    for table in tables:
        fingerprint = str(table.get("fingerprint", ""))
        if not fingerprint:
            continue
        profile = session.scalar(select(ParserProfile).where(ParserProfile.fingerprint == fingerprint))
        mapping = {
            "columns": table.get("columns", {}),
            "header_rows": table.get("header_rows", []),
            "raw_columns": table.get("raw_columns", {}),
        }
        if profile:
            profile.report_type = str(table.get("report_type", profile.report_type))
            profile.mapping = mapping
            profile.enabled = True
            continue
        session.add(
            ParserProfile(
                fingerprint=fingerprint,
                name=f"{table.get('sheet_name', 'Excel')} 表头映射",
                report_type=str(table.get("report_type", "unknown")),
                mapping=mapping,
            )
        )


def _apply_profile(session: Session, preview: dict[str, Any]) -> dict[str, Any]:
    fingerprint = str(preview.get("fingerprint", ""))
    if not fingerprint:
        return preview
    profile = session.scalar(
        select(ParserProfile).where(ParserProfile.fingerprint == fingerprint, ParserProfile.enabled.is_(True))
    )
    if not profile:
        return preview
    preview["report_type"] = profile.report_type
    preview["columns"] = profile.mapping.get("columns", preview["columns"])
    preview["raw_columns"] = profile.mapping.get("raw_columns", preview.get("raw_columns", {}))
    preview["requires_confirmation"] = False
    preview["profile_id"] = profile.id
    return preview


def _table_from_preview(data: dict[str, Any]) -> ParsedTable:
    columns = {
        field: int(details["column"])
        for field, details in data.get("columns", {}).items()
        if isinstance(details, dict) and details.get("column")
    }
    header_rows = [int(row) for row in data.get("header_rows", [])]
    header_paths = {
        field: list(details.get("header_path", []))
        for field, details in data.get("columns", {}).items()
        if isinstance(details, dict)
    }
    all_header_paths = {
        int(column_no): list(path)
        for column_no, path in data.get("raw_columns", {}).items()
        if str(column_no).isdigit() and isinstance(path, list)
    }
    return ParsedTable(
        report_type=str(data.get("report_type", "unknown")),
        sheet_name=str(data["sheet_name"]),
        header_row=max(header_rows),
        columns=columns,
        header_rows=header_rows,
        header_paths=header_paths,
        all_header_paths=all_header_paths,
    )


def _issue(
    session: Session,
    job_id: str,
    source_file_id: str | None,
    severity: str,
    code: str,
    message: str,
    sheet_name: str | None = None,
    cell_range: str | None = None,
    action: str | None = None,
) -> None:
    session.add(
        ImportIssue(
            import_job_id=job_id,
            source_file_id=source_file_id,
            severity=severity,
            code=code,
            message=message,
            sheet_name=sheet_name,
            cell_range=cell_range,
            suggested_action=action,
        )
    )


def _safe_scaled(value: Any, *, issue_context: tuple[Session, str, str, str, int], field: str) -> int | None:
    if value is None or value == "":
        return None
    try:
        return to_scaled(str(value))
    except ValueError:
        session, job_id, source_id, sheet, row_no = issue_context
        _issue(
            session,
            job_id,
            source_id,
            "warning",
            "INVALID_DECIMAL",
            f"{field} 无法解析为数值：{value}",
            sheet,
            str(row_no),
            "核对原始单元格或将该问题标记为忽略",
        )
        return None


def _resource_kind(source_category: str, name: str) -> tuple[str, bool]:
    """Prefer a mapped source category; only use a name heuristic as a reviewable fallback."""
    normalized = normalize_text(source_category)
    if any(word in normalized for word in ("人工", "劳务", "工日")):
        return "labor", True
    if any(word in normalized for word in ("机械", "机具", "台班")):
        return "machine", True
    if any(word in normalized for word in ("材料", "设备", "主材")):
        return "material", True
    lower_name = normalize_text(name)
    if "人工" in lower_name:
        return "labor", False
    if "机械" in lower_name or "台班" in lower_name:
        return "machine", False
    return "material", False


def _parse_table(
    session: Session,
    job: ImportJob,
    source_file: SourceFile,
    path: Path,
    table: ParsedTable,
    pending_analysis: list[dict[str, Any]],
) -> int:
    values_book = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    formula_book = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    count = 0
    try:
        values_sheet = values_book[table.sheet_name]
        formula_sheet = formula_book[table.sheet_name]
        hierarchy: list[str] = []
        for row_no, row in enumerate(
            values_sheet.iter_rows(min_row=table.header_row + 1, values_only=True),
            start=table.header_row + 1,
        ):
            name = _cell_text(_value(row, table.columns, "name"))
            code = _cell_text(_value(row, table.columns, "code")) or None
            if not name and not code:
                continue
            if not name:
                continue
            if name in {"合计", "小计", "总计"} or name.endswith("合计"):
                continue
            context = (session, job.id, source_file.id, table.sheet_name, row_no)
            unit = normalize_unit(_cell_text(_value(row, table.columns, "unit")), session) or None
            quantity = _safe_scaled(_value(row, table.columns, "quantity"), issue_context=context, field="工程量")
            unit_price = _safe_scaled(_value(row, table.columns, "unit_price"), issue_context=context, field="单价")
            total = _safe_scaled(_value(row, table.columns, "total"), issue_context=context, field="金额")

            # Preserve source structure instead of collapsing identically coded rows.
            if table.report_type in {"bill", "catalog_bill"} and not code and name and quantity is None and unit_price is None and total is None:
                hierarchy.append(name)
                continue

            for numeric_field in ("quantity", "unit_price", "total", "rate"):
                col = table.columns.get(numeric_field)
                if col:
                    formula_cell = formula_sheet.cell(row=row_no, column=col)
                    cached = _value(row, table.columns, numeric_field)
                    if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") and cached is None:
                        _issue(
                            session,
                            job.id,
                            source_file.id,
                            "warning",
                            "FORMULA_CACHE_MISSING",
                            f"公式单元格 {formula_cell.coordinate} 没有缓存结果",
                            table.sheet_name,
                            formula_cell.coordinate,
                            "使用 Excel 打开并重新保存后再次导入",
                        )

            if table.report_type in {"bill", "catalog_bill"}:
                attributes = {
                    field: _cell_text(_value(row, table.columns, field))
                    for field in table.columns
                    if field not in {"code", "name", "description", "specification", "unit", "quantity", "unit_price", "total"}
                    and _cell_text(_value(row, table.columns, field))
                }
                for column_no, header_path in table.all_header_paths.items():
                    if column_no in table.columns.values() or column_no > len(row):
                        continue
                    raw_value = _cell_text(row[column_no - 1])
                    if raw_value:
                        attributes.setdefault(" / ".join(header_path) or f"列{column_no}", raw_value)
                item = CostItem(
                    project_version_id=job.project_version_id,
                    source_file_id=source_file.id,
                    code=code,
                    name=name,
                    normalized_name=normalize_text(name),
                    description=_cell_text(_value(row, table.columns, "description")) or None,
                    specification=_cell_text(_value(row, table.columns, "specification")) or None,
                    unit=unit,
                    quantity_value=quantity,
                    unit_price_value=unit_price,
                    total_value=total,
                    sheet_name=table.sheet_name,
                    source_row=row_no,
                    source_cells={field: formula_sheet.cell(row=row_no, column=column_no).coordinate for field, column_no in table.columns.items()},
                    import_attributes=attributes,
                    hierarchy_path=" / ".join(hierarchy) or None,
                    data_status="parsed",
                    item_type="library_bill" if table.report_type == "catalog_bill" else "bill",
                )
                session.add(item)
            elif table.report_type == "resource":
                source_category = _cell_text(_value(row, table.columns, "category"))
                kind, confirmed = _resource_kind(source_category, name)
                if not confirmed:
                    _issue(session, job.id, source_file.id, "warning", "RESOURCE_KIND_UNCONFIRMED", f"{name} 未提供可确认的工料机类别，暂按{kind}处理", table.sheet_name, str(row_no), "确认来源类别后再用于分类统计")
                session.add(
                    ResourceItem(
                        project_version_id=job.project_version_id,
                        source_file_id=source_file.id,
                        kind=kind,
                        code=code,
                        name=name,
                        specification=_cell_text(_value(row, table.columns, "specification")) or None,
                        unit=unit,
                        quantity_value=quantity,
                        price_value=unit_price,
                        amount_value=total,
                        sheet_name=table.sheet_name,
                        source_row=row_no,
                        source_category=source_category or None,
                        source_cells={field: formula_sheet.cell(row=row_no, column=column_no).coordinate for field, column_no in table.columns.items()},
                        data_status="parsed",
                    )
                )
            elif table.report_type == "measure":
                session.add(
                    MeasureItem(
                        project_version_id=job.project_version_id,
                        source_file_id=source_file.id,
                        name=name,
                        amount_value=total if total is not None else unit_price,
                        sheet_name=table.sheet_name,
                        source_row=row_no,
                    )
                )
            elif table.report_type == "fee_rate":
                session.add(
                    FeeRate(
                        project_version_id=job.project_version_id,
                        source_file_id=source_file.id,
                        name=name,
                        rate_value=_safe_scaled(_value(row, table.columns, "rate"), issue_context=context, field="费率"),
                        basis=_cell_text(_value(row, table.columns, "basis")) or None,
                        sheet_name=table.sheet_name,
                        source_row=row_no,
                    )
                )
            elif table.report_type in {"rate_analysis", "quota"}:
                pending_analysis.append(
                    {
                        "type": table.report_type,
                        "source_file_id": source_file.id,
                        "sheet_name": table.sheet_name,
                        "source_row": row_no,
                        "code": code,
                        "name": name,
                        "category": _cell_text(_value(row, table.columns, "category")) or "其他",
                        "unit": unit,
                        "quantity": quantity,
                        "amount": total if total is not None else unit_price,
                    }
                )
            count += 1
    finally:
        values_book.close()
        formula_book.close()
    return count


def _link_analysis(session: Session, job: ImportJob, rows: list[dict[str, Any]]) -> None:
    session.flush()
    for row in rows:
        if not row["code"]:
            _issue(
                session,
                job.id,
                row["source_file_id"],
                "warning",
                "ANALYSIS_CODE_MISSING",
                f"{row['name']} 缺少关联清单编码",
                row["sheet_name"],
                str(row["source_row"]),
            )
            continue
        items = session.scalars(
            select(CostItem).where(
                CostItem.project_version_id == job.project_version_id,
                CostItem.code == row["code"],
            )
        ).all()
        if not items:
            _issue(
                session,
                job.id,
                row["source_file_id"],
                "warning",
                "ANALYSIS_LINK_FAILED",
                f"未找到编码 {row['code']} 对应的清单项",
                row["sheet_name"],
                str(row["source_row"]),
                "确认编码或在异常复核中忽略",
            )
            continue
        if len(items) > 1:
            _issue(
                session,
                job.id,
                row["source_file_id"],
                "warning",
                "ANALYSIS_LINK_AMBIGUOUS",
                f"编码 {row['code']} 匹配到 {len(items)} 条清单项，未自动关联 {row['name']}",
                row["sheet_name"],
                str(row["source_row"]),
                "在复核中按名称、单位、项目特征或父级路径确认关联",
            )
            continue
        item = items[0]
        evidence = {
            "strategy": "project_version + code",
            "candidate_count": 1,
            "candidate_cost_item_id": item.id,
            "analysis_source": {"file_id": row["source_file_id"], "sheet_name": row["sheet_name"], "row": row["source_row"]},
        }
        if row["type"] == "quota":
            session.add(
                QuotaItem(
                    cost_item_id=item.id,
                    code=row["code"],
                    name=row["name"],
                    unit=row["unit"],
                    consumption_value=row["quantity"],
                    source_file_id=row["source_file_id"],
                    sheet_name=row["sheet_name"],
                    source_row=row["source_row"],
                    link_status="confirmed",
                    link_evidence=evidence,
                )
            )
        else:
            session.add(
                RateComponent(
                    cost_item_id=item.id,
                    category=row["category"],
                    name=row["name"],
                    value=row["amount"],
                    source_file_id=row["source_file_id"],
                    sheet_name=row["sheet_name"],
                    source_row=row["source_row"],
                    link_status="confirmed",
                    link_evidence=evidence,
                )
            )


def process_import_job(job_id: str) -> None:
    settings = get_settings()
    with SessionLocal() as session:
        job = session.get(ImportJob, job_id)
        if not job:
            return
        job.status = "analyzing" if not job.parse_preview else "parsing"
        job.started_at = datetime.now(timezone.utc)
        session.commit()
        try:
            files = session.scalars(
                select(SourceFile).where(SourceFile.import_job_id == job.id).order_by(SourceFile.original_name)
            ).all()
            file_by_id = {source_file.id: source_file for source_file in files}
            if not job.parse_preview:
                preview_tables: list[dict[str, Any]] = []
                for source_file in files:
                    path = settings.raw_dir / source_file.relative_path
                    sheet_names, tables = inspect_workbook(path)
                    source_file.sheet_names = sheet_names
                    source_file.report_type = next((table.report_type for table in tables if table.report_type != "unknown"), "unknown")
                    preview_tables.extend(_apply_profile(session, _table_preview(source_file, table)) for table in tables)
                job.parse_preview = {"tables": preview_tables}
                if any(table["requires_confirmation"] for table in preview_tables):
                    job.status = "mapping_review"
                    job.progress = 20
                    session.commit()
                    return
            pending_analysis: list[dict[str, Any]] = []
            tables_by_file: dict[str, list[ParsedTable]] = {}
            for table_data in job.parse_preview.get("tables", []):
                source_file_id = str(table_data.get("source_file_id", ""))
                if source_file_id in file_by_id:
                    tables_by_file.setdefault(source_file_id, []).append(_table_from_preview(table_data))
            for position, source_file in enumerate(files, start=1):
                path = settings.raw_dir / source_file.relative_path
                try:
                    known_tables = [table for table in tables_by_file.get(source_file.id, []) if table.report_type != "unknown"]
                    if not known_tables:
                        _issue(
                            session,
                            job.id,
                            source_file.id,
                            "warning",
                            "REPORT_UNRECOGNIZED",
                            "未识别到受支持的报表表头",
                            action="确认表头映射或核对文件是否为受支持的报表",
                        )
                    for table in known_tables:
                        _parse_table(session, job, source_file, path, table, pending_analysis)
                except Exception as exc:
                    _issue(
                        session,
                        job.id,
                        source_file.id,
                        "error",
                        "WORKBOOK_READ_FAILED",
                        f"工作簿读取失败：{exc}",
                        action="确认文件未损坏且格式为 xlsx 或 xlsm",
                    )
                job.processed_files = position
                job.progress = int(position / max(len(files), 1) * 85)
                session.commit()
            _link_analysis(session, job, pending_analysis)
            item_count = session.scalar(
                select(func.count(CostItem.id)).where(CostItem.project_version_id == job.project_version_id)
            ) or 0
            if item_count == 0:
                _issue(
                    session,
                    job.id,
                    None,
                    "error",
                    "NO_COST_ITEMS",
                    "本次导入没有提取到任何清单项",
                    action="检查清单表表头或更换导入文件",
                )
            job.status = "review"
            job.progress = 100
            job.finished_at = datetime.now(timezone.utc)
            session.commit()
        except Exception as exc:
            session.rollback()
            job = session.get(ImportJob, job_id)
            if job:
                job.status = "failed"
                job.error_summary = str(exc)
                job.finished_at = datetime.now(timezone.utc)
                session.commit()


def next_version_no(session: Session, project_id: str) -> int:
    current = session.scalar(
        select(func.max(ProjectVersion.version_no)).where(ProjectVersion.project_id == project_id)
    )
    return (current or 0) + 1
