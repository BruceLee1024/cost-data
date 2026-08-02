from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from cost_data.config import get_settings
from cost_data.models import ImportIssue, Project, ProjectVersion, SourceFile
from cost_data.search import get_cost_item
from cost_data.libraries import LIBRARY_LABELS, get_record as get_library_record
from cost_data.governance import comparability, record_warnings


HEADER_FILL = PatternFill("solid", fgColor="173F35")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(sheet) -> None:  # type: ignore[no-untyped-def]
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"


def export_reference_prices(session: Session, item_ids: list[str], library: str | None = None) -> Path:
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "历史参考数据"
    sheet.append(
        [
            "数据来源库",
            "编码",
            "名称",
            "项目特征/附加属性",
            "规格",
            "单位",
            "原始单价",
            "原始金额",
            "项目名称",
            "地区",
            "计价时间",
            "专业",
            "计价体系",
            "价格口径",
            "可比性状态",
            "风险说明",
            "原始文件",
            "工作表",
            "来源行",
            "来源单元格",
        ]
    )
    _style_header(sheet)
    for item_id in item_ids:
        if library:
            row = get_library_record(library, item_id)
            version = session.get(ProjectVersion, row.project_version_id) if row else None
            project = session.get(Project, row.project_id) if row else None
            source = session.get(SourceFile, row.source_file_id) if row else None
            if not row or not version or version.status != "published" or not project:
                continue
            warnings = "；".join(record_warnings(project, link_status=row.payload.get("link_status"))) or "—"
            detail = row.payload.get("description") or row.payload.get("attributes") or "—"
            source_cells = row.payload.get("source_cells") or "—"
            sheet.append([LIBRARY_LABELS[library], row.code, row.name, json.dumps(detail, ensure_ascii=False) if isinstance(detail, dict) else detail, row.specification, row.unit, row.unit_price_value / (10 ** row.unit_price_scale) if row.unit_price_value is not None else None, row.total_value / (10 ** row.total_scale) if row.total_value is not None else None, project.name, project.region, project.pricing_date, project.specialty, project.pricing_mode, "；".join(f"{key}={value}" for key, value in project.price_context.items()) or "待确认", {"searchable": "检索可用", "restricted": "参考受限", "benchmark_candidate": "标杆候选"}[comparability(project)], warnings, source.original_name if source else "—", row.source_sheet, row.source_row, json.dumps(source_cells, ensure_ascii=False) if isinstance(source_cells, dict) else source_cells])
            continue
        item = get_cost_item(session, item_id)
        if item:
            sheet.append(["中心库", item.code, item.name, item.description, item.specification, item.unit, item.unit_price.value, item.total.value, item.project_name, item.region, item.pricing_date, item.specialty, item.pricing_mode, "待确认", "参考受限", "请从分库记录导出以取得完整价格口径", item.source.file_name, item.source.sheet_name, item.source.start_row, item.source.field_cells])
    widths = [14, 16, 32, 42, 24, 10, 16, 16, 28, 12, 14, 14, 16, 36, 14, 42, 28, 22, 10, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    path = get_settings().export_dir / "历史参考数据.xlsx"
    workbook.save(path)
    return path


def export_quality_report(session: Session, job_id: str) -> Path:
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "导入质量报告"
    sheet.append(["严重程度", "问题代码", "问题说明", "工作表", "位置", "状态", "建议处理"])
    _style_header(sheet)
    issues = session.scalars(
        select(ImportIssue).where(ImportIssue.import_job_id == job_id).order_by(ImportIssue.severity, ImportIssue.created_at)
    ).all()
    for issue in issues:
        sheet.append(
            [
                issue.severity,
                issue.code,
                issue.message,
                issue.sheet_name,
                issue.cell_range,
                issue.status,
                issue.suggested_action,
            ]
        )
    widths = [12, 24, 60, 24, 14, 12, 50]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    path = get_settings().export_dir / f"导入质量报告-{job_id[:8]}.xlsx"
    workbook.save(path)
    return path
