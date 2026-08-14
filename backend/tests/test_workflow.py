import json
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl import load_workbook

from cost_data.config import get_settings
from cost_data.main import app


def golden_workbook() -> bytes:
    workbook = Workbook()
    bill = workbook.active
    bill.title = "分部分项清单计价表"
    bill.append(["项目编码", "项目名称", "项目特征描述", "计量单位", "工程量", "综合单价", "合价"])
    bill.append(["010502001001", "现浇混凝土矩形柱", "混凝土强度等级 C30", "m3", 100, 685.25, 68525])
    bill.append(["010515001001", "现浇构件钢筋", "HRB400", "t", 12.5, 5320, 66500])
    resources = workbook.create_sheet("工料机汇总表")
    resources.append(["材料编码", "材料名称", "规格型号", "单位", "数量", "市场价", "金额"])
    resources.append(["MAT001", "商品混凝土", "C30", "m3", 105, 485, 50925])
    resources.append(["MAT002", "钢筋", "HRB400", "kg", 12500, 4.5, 56250])
    analysis = workbook.create_sheet("综合单价分析表")
    analysis.append(["项目编码", "项目名称", "费用类别", "金额"])
    analysis.append(["010502001001", "人工费", "人工", 65.25])
    analysis.append(["010502001001", "材料费", "材料", 580])
    measures = workbook.create_sheet("措施项目表")
    measures.append(["费用名称", "金额"])
    measures.append(["安全文明施工费", 12000])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def merged_header_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "分部分项工程量清单"
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "青岛住宅项目清单"
    sheet.merge_cells("A3:D3")
    sheet["A3"] = "清单项目"
    sheet.merge_cells("E3:G3")
    sheet["E3"] = "计价信息"
    sheet.append(["项目编码", "项目名称", "项目特征描述", "计量单位", "工程量", "综合单价", "合价"])
    sheet.append(["建筑工程", None, None, None, None, None, None])
    sheet.append(["010502001001", "现浇混凝土矩形柱", "混凝土强度等级 C30", "m3", 100, 685.25, 68525])
    sheet.append(["010515001001", "现浇构件钢筋", "HRB400", "t", 12.5, 5320, 66500])
    sheet.append([None, "分部合计", None, None, None, None, 135025])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def catalog_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清单库模板"
    sheet.append(["专业", "清单编码", "清单名称", "项目特征", "单位", "默认工程量", "适用范围", "结构分组", "关联定额编码", "来源", "版本", "备注", "业务标签"])
    sheet.append(["建筑工程", "010101001001", "平整场地", "项目特征：综合考虑", "㎡", 1133.76, "青岛住宅项目", "civil", "1-4-2", "原始清单.xlsx", "16建筑", "样本备注", "地下室"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def duplicate_code_analysis_workbook() -> bytes:
    workbook = Workbook()
    bill = workbook.active
    bill.title = "分部分项清单计价表"
    bill.append(["项目编码", "项目名称", "项目特征描述", "计量单位", "工程量", "综合单价", "合价"])
    bill.append(["010502001001", "现浇混凝土矩形柱", "C30", "m3", 10, 685.25, 6852.5])
    bill.append(["010502001001", "现浇混凝土异形柱", "C30", "m3", 8, 700, 5600])
    analysis = workbook.create_sheet("综合单价分析表")
    analysis.append(["项目编码", "项目名称", "费用类别", "金额"])
    analysis.append(["010502001001", "人工费", "人工", 65.25])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def governed_resource_workbook() -> bytes:
    workbook = Workbook()
    bill = workbook.active
    bill.title = "分部分项清单计价表"
    bill.append(["项目编码", "项目名称", "项目特征描述", "计量单位", "工程量", "综合单价", "合价"])
    bill.append(["010502001001", "现浇混凝土矩形柱", "混凝土强度等级 C30", "m3", 10, 685.25, 6852.5])
    resources = workbook.create_sheet("工料机汇总表")
    resources.append(["资源类别", "材料编码", "材料名称", "规格型号", "单位", "数量", "市场价", "金额"])
    resources.append(["机械", "MAC001", "混凝土泵送设备", "泵送", "台班", 2, 800, 1600])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def metadata(name: str = "西安人工黄金样本") -> str:
    return json.dumps({"name": name, "region": "西安", "pricing_date": "2026-06", "specialty": "土建", "pricing_mode": "内地清单定额", "result_stage": "结算", "area": "1000"}, ensure_ascii=False)


def test_import_publish_search_match_metrics_backup() -> None:
    client = TestClient(app)
    token = client.get("/api/v1/health").json()["session_token"]
    headers = {"X-Cost-Data-Token": token}
    workbook = golden_workbook()
    response = client.post("/api/v1/imports", headers=headers, data={"metadata_json": metadata()}, files=[("files", ("黄金样本.xlsx", workbook, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))])
    assert response.status_code == 202, response.text
    job = response.json()
    current = client.get(f"/api/v1/imports/{job['id']}").json()
    assert current["status"] == "review"
    assert current["processed_files"] == 1
    issues = client.get(f"/api/v1/imports/{job['id']}/issues").json()
    assert not [issue for issue in issues if issue["severity"] == "error"]

    published = client.post(f"/api/v1/imports/{job['id']}/publish", headers=headers)
    assert published.status_code == 200, published.text
    version_id = published.json()["id"]
    found = client.get("/api/v1/cost-items/search", params={"query": "混凝土"})
    assert found.status_code == 200, found.text
    assert found.json()["total"] == 1
    item = found.json()["items"][0]
    assert item["unit_price"]["value"] == "685.25"
    assert item["source"]["file_name"] == "黄金样本.xlsx"
    assert len(item["components"]) == 2
    quality = client.get(f"/api/v1/quality/projects/{version_id}")
    assert any(issue["code"] == "PRICE_CONTEXT_INCOMPLETE" for issue in quality.json()["issues"])

    match = client.post("/api/v1/match-sessions", headers=headers, json={"name": "测试匹配", "items": [{"name": "现浇混凝土柱", "unit": "m3"}]})
    assert match.status_code == 201, match.text
    assert match.json()["results"][0]["candidates"][0]["total_score"] > 0
    metrics = client.get(f"/api/v1/metrics/projects/{version_id}").json()
    assert {metric["code"] for metric in metrics} == {"cost_per_area", "steel_per_area", "concrete_per_area"}

    backup_dir = get_settings().data_home / "test-backups"
    backup = client.post("/api/v1/backups", headers=headers, json={"target_directory": str(backup_dir), "kind": "manual"})
    assert backup.status_code == 201, backup.text
    assert Path(backup.json()["path"], "manifest.json").exists()
    manifest = json.loads(Path(backup.json()["path"], "manifest.json").read_text(encoding="utf-8"))
    assert {entry["key"] for entry in manifest["databases"]} == {"center", "catalog", "resource", "quota"}

    duplicate = client.post("/api/v1/imports", headers=headers, data={"metadata_json": metadata(), "project_id": job["project_id"]}, files=[("files", ("黄金样本.xlsx", workbook, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))])
    assert duplicate.status_code == 409


def test_security_and_validation_contract() -> None:
    client = TestClient(app)
    denied = client.post("/api/v1/projects", json={})
    assert denied.status_code == 403
    response = client.post("/api/v1/projects", headers={"X-Cost-Data-Token": "test-session-token"}, json={})
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "VALIDATION_ERROR"
    vite_response = client.post(
        "/api/v1/projects",
        headers={"X-Cost-Data-Token": "test-session-token", "Origin": "http://127.0.0.1:5175"},
        json={},
    )
    assert vite_response.status_code == 422
    fallback_response = client.post(
        "/api/v1/projects",
        headers={"X-Cost-Data-Token": "test-session-token", "Origin": "http://127.0.0.1:8766"},
        json={},
    )
    assert fallback_response.status_code == 422
    external_origin = client.post(
        "/api/v1/projects",
        headers={"X-Cost-Data-Token": "test-session-token", "Origin": "https://example.com"},
        json={},
    )
    assert external_origin.status_code == 403
    preview = client.post(
        "/api/v1/ai/preview",
        headers={"X-Cost-Data-Token": "test-session-token"},
        json={"capability": "import_parsing", "payload": {"tables": []}},
    )
    assert preview.status_code == 200
    assert preview.json()["consent_required"] is True


def test_merged_headers_require_mapping_confirmation_before_publish() -> None:
    client = TestClient(app)
    token = client.get("/api/v1/health").json()["session_token"]
    headers = {"X-Cost-Data-Token": token}
    response = client.post(
        "/api/v1/imports",
        headers=headers,
        data={"metadata_json": metadata("青岛住宅一期")},
        files=[("files", ("多层表头.xlsx", merged_header_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
    )
    assert response.status_code == 202, response.text
    job = response.json()
    current = client.get(f"/api/v1/imports/{job['id']}").json()
    assert current["status"] == "mapping_review"

    preview = client.get(f"/api/v1/imports/{job['id']}/parse-preview")
    assert preview.status_code == 200, preview.text
    mapping = preview.json()["tables"][0]
    assert mapping["report_type"] == "bill"
    assert mapping["header_rows"] == [3, 4]
    assert mapping["columns"]["code"]["header_path"] == ["清单项目", "项目编码"]

    confirmed = client.post(
        f"/api/v1/imports/{job['id']}/confirm-mapping",
        headers=headers,
        json={"tables": [mapping], "save_profile": True},
    )
    assert confirmed.status_code == 200, confirmed.text
    current = client.get(f"/api/v1/imports/{job['id']}").json()
    assert current["status"] == "review"
    published = client.post(f"/api/v1/imports/{job['id']}/publish", headers=headers)
    assert published.status_code == 200, published.text
    items = client.get("/api/v1/cost-items/search", params={"query": "混凝土"}).json()
    assert any(item["source"]["file_name"] == "多层表头.xlsx" for item in items["items"])

    reused = client.post(
        "/api/v1/imports",
        headers=headers,
        data={"metadata_json": metadata("青岛住宅二期")},
        files=[("files", ("多层表头复用.xlsx", merged_header_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
    )
    assert reused.status_code == 202, reused.text
    reused_job = client.get(f"/api/v1/imports/{reused.json()['id']}").json()
    assert reused_job["status"] == "review"


def test_catalog_template_preserves_extra_fields_and_cell_evidence() -> None:
    client = TestClient(app)
    token = client.get("/api/v1/health").json()["session_token"]
    headers = {"X-Cost-Data-Token": token}
    response = client.post(
        "/api/v1/imports",
        headers=headers,
        data={"metadata_json": metadata("青岛清单库")},
        files=[("files", ("清单库样本.xlsx", catalog_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
    )
    assert response.status_code == 202, response.text
    job = response.json()
    assert client.get(f"/api/v1/imports/{job['id']}").json()["status"] == "review"
    assert client.post(f"/api/v1/imports/{job['id']}/publish", headers=headers).status_code == 200

    items = client.get("/api/v1/cost-items/search", params={"code": "010101001001"}).json()["items"]
    item = next(item for item in items if item["source"]["file_name"] == "清单库样本.xlsx")
    assert item["unit"] == "m2"
    assert item["quantity"]["value"] == "1133.76"
    assert item["item_type"] == "library_bill"
    assert item["import_attributes"]["structure_group"] == "civil"
    assert item["import_attributes"]["业务标签"] == "地下室"
    assert item["source"]["field_cells"]["code"] == "B2"


def test_duplicate_codes_do_not_create_ambiguous_analysis_link() -> None:
    client = TestClient(app)
    token = client.get("/api/v1/health").json()["session_token"]
    headers = {"X-Cost-Data-Token": token}
    response = client.post(
        "/api/v1/imports",
        headers=headers,
        data={"metadata_json": metadata("重复编码测试")},
        files=[("files", ("重复编码.xlsx", duplicate_code_analysis_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
    )
    assert response.status_code == 202, response.text
    issues = client.get(f"/api/v1/imports/{response.json()['id']}/issues").json()
    assert any(issue["code"] == "ANALYSIS_LINK_AMBIGUOUS" for issue in issues)


def test_phase_one_workspace_profile_quality_and_master_data() -> None:
    client = TestClient(app)
    token = client.get("/api/v1/health").json()["session_token"]
    headers = {"X-Cost-Data-Token": token}
    response = client.post(
        "/api/v1/imports",
        headers=headers,
        data={"metadata_json": json.dumps({"name": "首期闭环样本", "region": "西安", "pricing_date": "2026-06", "specialty": "土建", "pricing_mode": "内地清单定额", "result_stage": "结算", "project_type": "住宅", "area": "1000", "price_context": {"tax_inclusion": "含税", "price_type": "结算价", "price_source": "广联达导出"}}, ensure_ascii=False)},
        files=[("files", ("首期.xlsx", golden_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
    )
    job = response.json()
    published = client.post(f"/api/v1/imports/{job['id']}/publish", headers=headers).json()
    project_id = job["project_id"]
    updated = client.patch(
        f"/api/v1/projects/{project_id}/profile",
        headers=headers,
        json={"profile": {"structure_form": "框架", "area_basis": "总建筑面积", "above_ground_area": "900", "underground_area": "100"}},
    )
    assert updated.status_code == 200
    assert updated.json()["comparability"] == "benchmark_candidate"
    detail = client.get(f"/api/v1/projects/{project_id}")
    assert detail.status_code == 200
    assert detail.json()["data_counts"]["resource"] == 2
    workspace = client.get("/api/v1/workspace/search", params={"data_type": "resource", "query": "钢筋"})
    assert workspace.status_code == 200
    assert workspace.json()["items"][0]["data_type"] == "resource"
    assert workspace.json()["items"][0]["comparability"] == "benchmark_candidate"
    all_workspace = client.get("/api/v1/workspace/search", params={"data_type": "all"})
    assert all_workspace.status_code == 200
    assert {item["data_type"] for item in all_workspace.json()["items"]} >= {"bill", "resource", "measure", "metric"}
    assert {item["library"] for item in all_workspace.json()["items"] if item["data_type"] in {"bill", "resource"}} == {"catalog", "resource"}
    libraries = client.get("/api/v1/libraries")
    assert libraries.status_code == 200
    library_data = {entry["key"]: entry for entry in libraries.json()}
    assert library_data["catalog"]["record_count"] >= 2
    assert library_data["resource"]["record_count"] >= 2
    assert (get_settings().database_dir / "catalog.sqlite3").exists()
    assert (get_settings().database_dir / "resource.sqlite3").exists()
    catalog = client.get("/api/v1/libraries/catalog/search", params={"query": "混凝土"})
    assert catalog.status_code == 200
    catalog_item = catalog.json()["items"][0]
    assert catalog_item["library"] == "catalog"
    reference_search = client.get("/api/v1/libraries/catalog/search", params={"query": "C30", "reference_scope": "available", "specialty": "土建", "pricing_date_from": "2026-01", "pricing_date_to": "2026-12", "tax_inclusion": "含税", "price_min": "600", "sort_by": "unit_price"})
    assert reference_search.status_code == 200, reference_search.text
    assert reference_search.json()["total"] == 1
    assert reference_search.json()["available_count"] == 1
    assert reference_search.json()["restricted_count"] == 0
    assert reference_search.json()["items"][0]["description"] == "混凝土强度等级 C30"
    detail = client.get(f"/api/v1/libraries/catalog/records/{catalog_item['id']}")
    assert detail.status_code == 200
    assert detail.json()["source"]["file_name"] == "首期.xlsx"
    quality = client.get(f"/api/v1/quality/projects/{published['id']}")
    assert quality.status_code == 200
    assert quality.json()["publishable"] is True
    benchmark = client.get("/api/v1/benchmarks/metrics/steel_per_area", params={"project_type": "住宅", "specialty": "土建"})
    assert benchmark.status_code == 200
    assert benchmark.json()["sample_count"] == 1
    conversion = client.post("/api/v1/unit-conversions", headers=headers, json={"source_unit": "箱", "target_unit": "个", "factor": "12", "basis": "包装规格"})
    assert conversion.status_code == 201
    template = client.post("/api/v1/metric-templates", headers=headers, json={"code": "wall_area", "name": "墙体面积", "unit": "m2", "formula": "墙体工程量汇总 ÷ 建筑面积"})
    assert template.status_code == 201


def test_library_projection_keeps_raw_price_context_category_and_export_evidence() -> None:
    client = TestClient(app)
    headers = {"X-Cost-Data-Token": client.get("/api/v1/health").json()["session_token"]}
    governed_metadata = json.dumps(
        {
            "name": "治理样本",
            "region": "西安",
            "pricing_date": "2026-06",
            "specialty": "土建",
            "pricing_mode": "内地清单定额",
            "result_stage": "结算",
            "area": "1000",
            "price_context": {"tax_inclusion": "含税", "price_type": "结算价", "price_source": "工料机汇总表"},
        },
        ensure_ascii=False,
    )
    created = client.post(
        "/api/v1/imports",
        headers=headers,
        data={"metadata_json": governed_metadata},
        files=[("files", ("治理样本.xlsx", governed_resource_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
    )
    assert created.status_code == 202, created.text
    job = created.json()
    assert client.post(f"/api/v1/imports/{job['id']}/publish", headers=headers).status_code == 200

    resource = client.get("/api/v1/libraries/resource/search", params={"data_status": "published", "resource_kind": "machine"})
    assert resource.status_code == 200, resource.text
    record = resource.json()["items"][0]
    assert record["data_status"] == "published"
    assert record["attributes"]["kind"] == "machine"
    assert record["attributes"]["source_cells"]["unit_price"] == "G2"
    assert record["price_context"]["tax_inclusion"] == "含税"
    assert record["comparability"] == "restricted"

    exported = client.get("/api/v1/exports/reference-prices", params=[("ids", record["id"]), ("library", "resource")])
    assert exported.status_code == 200, exported.text
    export_path = get_settings().export_dir / "历史参考数据.xlsx"
    workbook = load_workbook(export_path, data_only=True)
    sheet = workbook.active
    headers_in_export = [cell.value for cell in sheet[1]]
    assert {"数据来源库", "价格口径", "可比性状态", "风险说明", "原始文件", "来源单元格"}.issubset(headers_in_export)
    values = [cell.value for cell in sheet[2]]
    assert "工料机库" in values
    assert any("含税" in str(value) for value in values)
    assert "参考受限" in values
