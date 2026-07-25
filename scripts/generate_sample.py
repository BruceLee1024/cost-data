from pathlib import Path

from openpyxl import Workbook


def main() -> None:
    target = Path(__file__).resolve().parents[1] / "tests" / "fixtures" / "synthetic-glodon-style.xlsx"
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "分部分项清单计价表"
    sheet.append(["项目编码", "项目名称", "项目特征描述", "计量单位", "工程量", "综合单价", "合价"])
    sheet.append(["010502001001", "现浇混凝土矩形柱", "混凝土强度等级 C30", "m3", 100, 685.25, 68525])
    sheet = workbook.create_sheet("工料机汇总表")
    sheet.append(["材料编码", "材料名称", "规格型号", "单位", "数量", "市场价", "金额"])
    sheet.append(["MAT001", "商品混凝土", "C30", "m3", 105, 485, 50925])
    sheet = workbook.create_sheet("综合单价分析表")
    sheet.append(["项目编码", "项目名称", "费用类别", "金额"])
    sheet.append(["010502001001", "人工费", "人工", 65.25])
    workbook.save(target)
    print(target)


if __name__ == "__main__":
    main()
